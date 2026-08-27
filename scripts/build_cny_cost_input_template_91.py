#!/usr/bin/env python3
"""Create zero-write CNY cost input templates for the 91 currently unresolved products."""
from __future__ import annotations

import csv
import json
import os
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "reports"
RECON = REPORT_DIR / "fx_pricing_local_source_reconciliation_2026-08-27.json"
GAP = REPORT_DIR / "fx_pricing_metadata_gap_audit_2026-08-27.json"
PRODUCT_CSV = REPORT_DIR / "cny_cost_input_product_list_91_2026-08-27.csv"
PRICE_CSV = REPORT_DIR / "cny_cost_input_by_active_price_91_2026-08-27.csv"
XLSX = REPORT_DIR / "cny_cost_input_template_91_2026-08-27.xlsx"
SUMMARY = REPORT_DIR / "cny_cost_input_template_91_2026-08-27.md"


def text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def api_get(api_key: str, path: str, params: dict[str, str]) -> dict[str, Any]:
    response = requests.get(f"https://api.stripe.com/v1/{path}", params=params, auth=(api_key, ""), timeout=60)
    response.raise_for_status()
    return response.json()


def list_all(api_key: str, path: str, params: dict[str, str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    cursor: str | None = None
    while True:
        page_params = dict(params)
        if cursor:
            page_params["starting_after"] = cursor
        page = api_get(api_key, path, page_params)
        records.extend(page["data"])
        if not page.get("has_more"):
            return records
        cursor = page["data"][-1]["id"]
        if not cursor:
            raise RuntimeError(f"Incomplete Stripe pagination for {path}")


def product_id_for_price(price: dict[str, Any]) -> str:
    product = price.get("product")
    return product if isinstance(product, str) else text((product or {}).get("id"))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]) if rows else [], lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="5B3A29")
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row_index, col, row.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, header in enumerate(headers, start=1):
        width = min(max(12, len(header) + 4, *(len(str(row.get(header, ""))) for row in rows[:50])), 48)
        worksheet.column_dimensions[get_column_letter(col)].width = width
    worksheet.auto_filter.ref = worksheet.dimensions


def main() -> None:
    api_key = os.environ.get("STRIPE_SECRET_KEY")
    if not api_key:
        raise SystemExit("STRIPE_SECRET_KEY is required")
    recon = json.loads(RECON.read_text(encoding="utf-8"))
    gap = json.loads(GAP.read_text(encoding="utf-8"))
    gap_by_product = {item["product_id"]: item for item in gap["excluded_products"]}
    unresolved = [item for item in recon["products"] if not item["local_cost_records"]]
    if len(unresolved) != 91:
        raise RuntimeError(f"Expected exactly 91 unresolved products, found {len(unresolved)}")

    products = list_all(api_key, "products", {"active": "true", "limit": "100"})
    prices = list_all(api_key, "prices", {"active": "true", "currency": "hkd", "limit": "100"})
    products_by_id = {product["id"]: product for product in products}
    prices_by_product: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for price in prices:
        product_id = product_id_for_price(price)
        if product_id:
            prices_by_product[product_id].append(price)

    product_rows: list[dict[str, str]] = []
    price_rows: list[dict[str, str]] = []
    errors: list[str] = []
    for number, record in enumerate(sorted(unresolved, key=lambda item: (item["mofu_sku"], item["product_id"])), start=1):
        product_id = record["product_id"]
        product = products_by_id.get(product_id)
        if not product:
            errors.append(f"Active Stripe Product no longer found: {product_id}")
            continue
        product_md = product.get("metadata") or {}
        audit = gap_by_product.get(product_id)
        if not audit:
            errors.append(f"Metadata audit record missing: {product_id}")
            continue
        active_prices = sorted(prices_by_product.get(product_id, []), key=lambda item: item["id"])
        if not active_prices:
            errors.append(f"Active HKD Price missing: {product_id}")
            continue
        product_rows.append({
            "序號": str(number),
            "店內 SKU": text(product_md.get("mofu_sku")),
            "產品名稱": text(product.get("name")),
            "品牌": text(product_md.get("brand")),
            "分類": text(product_md.get("category")),
            "供應商 SKU": text(product_md.get("sku")),
            "Stripe Product ID": product_id,
            "活躍 HKD Price 數": str(len(active_prices)),
            "現行 HKD 售價／Price ID／規格": " | ".join(
                f"HK${(price.get('unit_amount') or 0) / 100:.2f} · {price['id']} · {text((price.get('metadata') or {}).get('variant_label_zh')) or '單一規格'}"
                for price in active_prices
            ),
            "原始匯入來源": record.get("mofu_import_source", ""),
            "原始匯入鍵": record.get("mofu_import_key", ""),
            "現有 Product metadata 欄位數": str(len(product_md)),
            "需先作欄位容量審核": "是" if len(product_md) >= 50 else "否",
            "待輸入 CNY 成本（如各變體相同）": "",
            "成本來源／供應商連結或單據": "",
            "成本日期": "",
            "確認目標毛利率（例如 0.45）": "",
            "入境／運費 HKD（如適用）": "",
            "備註": "",
        })
        for price in active_prices:
            price_md = price.get("metadata") or {}
            price_rows.append({
                "產品序號": str(number),
                "店內 SKU": text(product_md.get("mofu_sku")),
                "產品名稱": text(product.get("name")),
                "品牌": text(product_md.get("brand")),
                "供應商 SKU": text(product_md.get("sku")),
                "Stripe Product ID": product_id,
                "Stripe Price ID": price["id"],
                "Price 是否預設": "是" if product.get("default_price") == price["id"] else "否",
                "變體／規格": text(price_md.get("variant_label_zh")) or "單一規格",
                "現行 HKD 售價": f"{(price.get('unit_amount') or 0) / 100:.2f}",
                "原始匯入來源": record.get("mofu_import_source", ""),
                "待輸入 CNY 成本（此 Price）": "",
                "成本來源／供應商連結或單據": "",
                "成本日期": "",
                "確認目標毛利率（例如 0.45）": "",
                "入境／運費 HKD（如適用）": "",
                "備註": "",
            })
    product_skus = [row["店內 SKU"] for row in product_rows]
    price_ids = [row["Stripe Price ID"] for row in price_rows]
    if errors or len(product_rows) != 91 or len(set(product_skus)) != 91 or len(price_ids) != len(set(price_ids)):
        raise RuntimeError("; ".join(errors) or "Product/SKU/Price ID template integrity check failed")

    write_csv(PRODUCT_CSV, product_rows)
    write_csv(PRICE_CSV, price_rows)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "填寫說明"
    readme.append(["91 件未有 CNY 成本產品 — 填寫說明"])
    readme["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor="5B3A29")
    instructions = [
        "本檔只供補充成本資料；不會自動更新 Stripe 或售價。",
        "如同一產品每個變體成本相同，可在「逐產品概覽」填寫；如不同，請在「逐 Price 輸入」逐行填寫。",
        "CNY 成本應為單件／該變體的實際進貨成本；請同時保留供應商連結、單據或其他可核對來源。",
        "目標毛利率及入境／運費必須由店主確認，切勿假設。",
        "標示「需先作欄位容量審核＝是」的產品，即使找回成本，也需另行清理 Product metadata 後才可同步。",
        "填妥後請把檔案回傳；下一輪會先做唯讀驗證及重定價預覽，未經明確批准不會改價。",
    ]
    for line in instructions:
        readme.append([line])
    readme.column_dimensions["A"].width = 120
    for row in range(2, len(instructions) + 2):
        readme.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    product_headers = list(product_rows[0])
    price_headers = list(price_rows[0])
    add_sheet(workbook, "逐產品概覽（91 件）", product_headers, product_rows)
    add_sheet(workbook, "逐 Price 輸入", price_headers, price_rows)
    workbook.save(XLSX)
    verified_workbook = load_workbook(XLSX, read_only=True)
    if (
        verified_workbook.sheetnames != ["填寫說明", "逐產品概覽（91 件）", "逐 Price 輸入"]
        or verified_workbook["逐產品概覽（91 件）"].max_row != 92
        or verified_workbook["逐 Price 輸入"].max_row != len(price_rows) + 1
    ):
        raise RuntimeError("Excel workbook integrity check failed")

    source_counts = Counter(row["原始匯入來源"] for row in product_rows)
    capacity_review_count = sum(row["需先作欄位容量審核"] == "是" for row in product_rows)
    summary = f"""# 91 件產品 CNY 成本輸入清單

**產生時間（UTC）：** {datetime.now(timezone.utc).isoformat()}
**狀態：** 唯讀清單；本次只重新讀取 Stripe 現行 Product／HKD Price 以建立輸入模板，**沒有寫入 metadata 或修改售價。**

本清單有 **91 件產品**及其 **{len(price_rows)} 項活躍 HKD Price**。Excel 檔包含「逐產品概覽」及「逐 Price 輸入」兩個工作表；如產品有不同規格或價格，請優先在「逐 Price 輸入」填寫該行的 CNY 成本。

| 提示 | 數量 |
| --- | ---: |
| 未有本機可核對 CNY 成本來源的產品 | 91 |
| 對應的活躍 HKD Price | {len(price_rows)} |
| Product metadata 已滿 50 欄、日後需先處理容量 | {capacity_review_count} |
| 已改動 Stripe metadata／Price | 0 |

| 原始匯入來源 | 產品數 |
| --- | ---: |
{chr(10).join(f'| {source} | {count} |' for source, count in sorted(source_counts.items()))}

請填寫每個成本的 CNY 金額、來源／單據、成本日期，並確認目標毛利率與任何入境／運費。填妥後回傳 Excel 或 CSV；我會先生成唯讀驗證及重定價預覽，不會直接改價。
"""
    SUMMARY.write_text(summary, encoding="utf-8")
    print(json.dumps({
        "mode": "read_only_cost_input_template",
        "no_stripe_writes_performed": True,
        "product_count": len(product_rows),
        "active_hkd_price_count": len(price_rows),
        "metadata_capacity_review_product_count": capacity_review_count,
        "source_counts": dict(sorted(source_counts.items())),
        "product_csv": str(PRODUCT_CSV),
        "price_csv": str(PRICE_CSV),
        "xlsx": str(XLSX),
        "summary": str(SUMMARY),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
