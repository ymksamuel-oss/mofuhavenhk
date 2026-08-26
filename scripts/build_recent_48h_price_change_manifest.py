from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/home/ubuntu/mofuhavenhk-github')
retail_audit = json.loads((ROOT / 'recent_48h_retail_formula_audit.json').read_text(encoding='utf-8'))
compare_candidates = json.loads((ROOT / 'recent_48h_compare_at_candidates.json').read_text(encoding='utf-8'))
pricing_input = json.loads((ROOT / 'recent_48h_pricing_audit_input.json').read_text(encoding='utf-8'))

existing_compare = {row['product_id']: row['existing_compare_at_hkd'] for row in pricing_input['products']}
external_candidate_ids = {item['product_id'] for item in compare_candidates['candidates']}

retail_updates = [
    row for row in retail_audit['rows']
    if row['pricing_status'] == 'recalculable' and float(row['price_delta_hkd']) != 0
]
retail_unchanged = [
    row for row in retail_audit['rows']
    if row['pricing_status'] == 'recalculable' and float(row['price_delta_hkd']) == 0
]
missing_cost = [row for row in retail_audit['rows'] if row['pricing_status'] != 'recalculable']

for row in retail_updates:
    row['retail_update_action'] = 'create_new_hkd_price_and_retire_old_active_price'
    row['reason'] = 'Recalculated at 45% product gross-margin target and rounded up to next .90 price point using recorded landed cost.'
for row in retail_unchanged:
    row['retail_update_action'] = 'no_change'
    row['reason'] = 'Current retail already equals calculated .90 target price.'
for row in missing_cost:
    row['retail_update_action'] = 'no_change'
    row['reason'] = 'No confirmed cost record; do not estimate or change retail price.'

# Existing compare-at stays as-is. External competitor / marketplace references are documented but
# withheld from the storefront "原價" field because they are not Mofu Haven historical prices.
compare_updates: list[dict] = []
market_reference_withheld = []
for record in compare_candidates['candidates']:
    record = dict(record)
    record['action'] = 'withhold_from_compare_at'
    record['reason'] = (
        'External retailer original/list price is a market reference, not verified Mofu Haven historical price. '
        'Do not map it to storefront 原價 without a dedicated market-reference label and merchant approval.'
    )
    if record['product_id'] in {'prod_V8e143elQIOfMM', 'prod_V8e2wynYC6XsBo'}:
        record['reason'] += ' Potential duplicate D1104 SKU requires package/SKU reconciliation first.'
    market_reference_withheld.append(record)

manifest = {
    'schema': 'recent-48h-pricing-change-v1',
    'pricing_policy': retail_audit['policy'],
    'scope': {
        'total_products': retail_audit['total_products'],
        'recalculable_products': retail_audit['recalculable_products'],
        'retail_price_updates': len(retail_updates),
        'already_at_formula_price': len(retail_unchanged),
        'cost_data_missing': len(missing_cost),
        'existing_compare_at_unchanged': sum(1 for value in existing_compare.values() if value not in ('', '0', '0.0', '0.00')),
        'new_compare_at_updates': 0,
        'external_market_references_withheld': len(market_reference_withheld),
    },
    'retail_updates': retail_updates,
    'retail_unchanged': retail_unchanged,
    'cost_missing': missing_cost,
    'existing_compare_at': [
        {'product_id': product_id, 'compare_at_hkd': value}
        for product_id, value in existing_compare.items()
        if value not in ('', '0', '0.0', '0.00')
    ],
    'market_reference_withheld': market_reference_withheld,
}
(ROOT / 'recent_48h_price_change_manifest.json').write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

# Human-review workbook
wb = Workbook()
ws = wb.active
ws.title = '摘要'
summary_rows = [
    ['項目', '數量'],
    ['最近兩日新增產品總數', manifest['scope']['total_products']],
    ['有可追溯成本，可按同一公式重算', manifest['scope']['recalculable_products']],
    ['建議建立新 Stripe 價格的產品', manifest['scope']['retail_price_updates']],
    ['現價已符合公式的產品', manifest['scope']['already_at_formula_price']],
    ['成本紀錄未完整，維持現價', manifest['scope']['cost_data_missing']],
    ['現有劃線價，保持不變', manifest['scope']['existing_compare_at_unchanged']],
    ['新增劃線價更新', manifest['scope']['new_compare_at_updates']],
    ['外部市場參考價，保留作研究而不當作本店原價', manifest['scope']['external_market_references_withheld']],
]
for row in summary_rows:
    ws.append(row)
ws['A1'].font = ws['B1'].font = Font(bold=True, color='FFFFFF')
ws['A1'].fill = ws['B1'].fill = PatternFill('solid', fgColor='6B4F3F')
ws.column_dimensions['A'].width = 48
ws.column_dimensions['B'].width = 18


def populate_sheet(title: str, rows: list[dict], headers: list[str]) -> None:
    sheet = wb.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='6B4F3F')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in rows:
        sheet.append([row.get(header, '') for header in headers])
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, 1):
        width = 18
        if header in {'name', 'reason', 'cost_source', 'evidence', 'source_url'}:
            width = 42
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    return sheet

price_headers = ['product_id','sku','family','name','current_retail_hkd','landed_cost_hkd','target_price_45_raw_hkd','target_charm_price_hkd','current_gross_margin_pct','price_delta_hkd','cost_source','retail_update_action','reason']
populate_sheet('零售價更新建議', retail_updates, price_headers)
populate_sheet('成本缺口，維持現價', missing_cost, ['product_id','sku','family','name','current_retail_hkd','cost_source','reason'])
populate_sheet('市場參考價（不採用）', market_reference_withheld, ['product_id','family','product_name','mofu_retail_hkd','source_current_price_hkd','source_original_price_hkd','source_retailer','source_url','reason'])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 34
    if sheet.max_row > 1:
        sheet.conditional_formatting.add(f'E2:E{sheet.max_row}', ColorScaleRule(start_type='min', start_color='F7E8E5', mid_type='percentile', mid_value=50, mid_color='F5D5A2', end_type='max', end_color='9CC9B3'))

wb.save(ROOT / 'MofuHaven_最近兩日產品定價與劃線價稽核.xlsx')
print(json.dumps(manifest['scope'], ensure_ascii=False))
