from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/mofuhavenhk-github')
XLSX = Path('/home/ubuntu/upload/dbf_dog_cans_23_items(1).xlsx')
UPLOAD_LOG = ROOT / 'dbf_clean_cdn_upload.json'
OUTPUT = ROOT / 'dbf_dog_cans_mapping.json'

IMAGES = {
    'DBF-01': 'dbf-01-chicken-cartilage-85g.png',
    'DBF-02': 'dbf-02-chicken-cheese-85g.png',
    'DBF-03': 'dbf-03-chicken-soup-85g.png',
    'DBF-04': 'dbf-04-chicken-cheese-adult-85g.png',
    'DBF-05': 'dbf-05-chicken-vegetables-85g.png',
    'DBF-06': 'dbf-06-beef-cartilage-85g.png',
    'DBF-07': 'dbf-07-beef-mince-65g.png',
    'DBF-08': 'dbf-08-chicken-sweetpotato-85g.png',
    'DBF-09': 'dbf-09-chicken-gizzard-soup-85g.png',
    'DBF-10': 'dbf-10-chicken-mince-65g.png',
    'DBF-11': 'dbf-11-adult-chicken-cartilage-85g.png',
    'DBF-12': 'dbf-12-adult-chicken-sweetpotato-85g.png',
    'DBF-13': 'dbf-13-adult-chicken-vegetables-85g.png',
    'DBF-14': 'dbf-14-senior-chicken-sweetpotato-85g.png',
    'DBF-15': 'dbf-15-beef-sweetpotato-85g.png',
    'DBF-16': 'dbf-16-senior-chicken-85g.png',
    'DBF-17': 'dbf-17-senior-chicken-vegetables-85g.png',
    'DBF-18': 'dbf-18-adult-chicken-85g.png',
    'DBF-19': 'dbf-19-beef-cheese-85g.png',
    'DBF-20': 'dbf-20-senior-chicken-cartilage-85g.png',
    'DBF-21': 'dbf-21-chicken-breast-mince-lowfat-65g.png',
    'DBF-22': 'dbf-22-chicken-chicken-breast-vegetables-150g.png',
    'DBF-23': 'dbf-23-chicken-chicken-breast-sweetpotato-150g.png',
}

ENGLISH_NAMES = {
    'DBF-01': 'd.b.f Chicken Breast & Cartilage Dog Can 85g',
    'DBF-02': 'd.b.f Chicken Breast & Cheese Dog Can 85g',
    'DBF-03': 'd.b.f Chicken Breast in Soup Dog Can 85g',
    'DBF-04': 'd.b.f Chicken & Cheese Adult Dog Can 85g',
    'DBF-05': 'd.b.f Chicken & Vegetables Adult Dog Can 85g',
    'DBF-06': 'd.b.f Beef & Cartilage Dog Can 85g',
    'DBF-07': 'd.b.f Beef Mince Dog Can 65g',
    'DBF-08': 'd.b.f Chicken & Sweet Potato Adult Dog Can 85g',
    'DBF-09': 'd.b.f Chicken Breast & Gizzard Soup Dog Can 85g',
    'DBF-10': 'd.b.f Chicken Mince Dog Can 65g',
    'DBF-11': 'd.b.f Adult Dog Meal Chicken & Cartilage 85g',
    'DBF-12': 'd.b.f Adult Dog Meal Chicken & Sweet Potato 85g',
    'DBF-13': 'd.b.f Adult Dog Meal Chicken & Vegetables 85g',
    'DBF-14': 'd.b.f Senior Dog Meal Chicken & Sweet Potato 85g',
    'DBF-15': 'd.b.f Beef & Sweet Potato Dog Can 85g',
    'DBF-16': 'd.b.f Senior Dog Meal Chicken 85g',
    'DBF-17': 'd.b.f Senior Dog Meal Chicken & Vegetables 85g',
    'DBF-18': 'd.b.f Adult Dog Meal Chicken 85g',
    'DBF-19': 'd.b.f Beef & Cheese Dog Can 85g',
    'DBF-20': 'd.b.f Senior Dog Meal Chicken & Cartilage 85g',
    'DBF-21': 'd.b.f Chicken Breast Mince Low-fat Dog Can 65g',
    'DBF-22': 'd.b.f Chicken & Chicken Breast Mince with Vegetables 150g',
    'DBF-23': 'd.b.f Chicken & Chicken Breast Mince with Sweet Potato 150g',
}


workbook = load_workbook(XLSX, data_only=True)
worksheet = workbook.active
headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
expected_headers = ['產品編號', '產品名稱 (中文)', '日文原名', '規格', '人民幣成本 (¥)', '建議零售價 (HKD)', '庫存狀態']
if headers != expected_headers:
    raise RuntimeError(f'Unexpected workbook headers: {headers!r}')

log_text = UPLOAD_LOG.read_text(encoding='utf-8')
cdn_by_filename = dict(re.findall(r'\[SUCCESS\] assets/dbf-dog-cans-clean/([^ ]+) -> (https://files\.manuscdn\.com/\S+)', log_text))
if len(cdn_by_filename) != 23:
    raise RuntimeError(f'Expected 23 cleaned CDN URLs, found {len(cdn_by_filename)}')

products = []
for row in worksheet.iter_rows(min_row=2, values_only=True):
    code, name_zh, japanese_name, spec, rmb_cost, retail_hkd, stock_status = row
    if code not in IMAGES or code not in ENGLISH_NAMES:
        raise RuntimeError(f'Missing local image or English name mapping for {code}')
    image = IMAGES[code]
    cdn_url = cdn_by_filename.get(image)
    if not cdn_url:
        raise RuntimeError(f'Missing CDN URL for {code}: {image}')
    if stock_status not in {'現貨', '缺貨'}:
        raise RuntimeError(f'Unexpected stock status for {code}: {stock_status!r}')
    products.append({
        'sku': code,
        'name_zh': name_zh,
        'name_en': ENGLISH_NAMES[code],
        'japanese_name': japanese_name if japanese_name != '-' else '',
        'spec': spec,
        'retail_hkd': float(retail_hkd),
        'compare_at_hkd': 0,
        'stock_status_zh': stock_status,
        'in_stock': stock_status == '現貨',
        'image_file': image,
        'cdn_url': cdn_url,
        'category': 'dogs',
        'subcategory': '狗狗食品',
        'product_type': 'dog_canned_food',
    })

if len(products) != 23:
    raise RuntimeError(f'Expected 23 products, got {len(products)}')
if [p['sku'] for p in products] != [f'DBF-{n:02d}' for n in range(1, 24)]:
    raise RuntimeError('Product SKU sequence is incomplete or duplicated')

output = {
    'source_document': XLSX.name,
    'brand': 'd.b.f',
    'product_count': len(products),
    'category': 'dogs',
    'subcategory': '狗狗食品',
    'out_of_stock_skus': [p['sku'] for p in products if not p['in_stock']],
    'products': products,
}
OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'output': str(OUTPUT),
    'product_count': len(products),
    'in_stock_count': sum(p['in_stock'] for p in products),
    'out_of_stock_skus': output['out_of_stock_skus'],
}, ensure_ascii=False))
